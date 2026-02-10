from fastapi import FastAPI, UploadFile, File, HTTPException
from fastapi.responses import StreamingResponse
from fastapi.middleware.cors import CORSMiddleware
import pandas as pd
import io
import json
from main import transform_data, load_mapping
from config import Config

app = FastAPI()

# Configurar CORS para permitir peticiones desde el frontend (React)
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"], # Allow ALL origins to fix local CORS issues (e.g. if port changes)
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
    expose_headers=["Content-Disposition"] # Required for frontend to read filename
)

@app.get("/")
def read_root():
    return {"status": "online", "message": "ETL API is running. Use POST /unify to upload files."}

def sanitize_cols(df):
    df.columns = [str(c).strip().replace(' ', '_').replace('.', '_').upper() for c in df.columns]
    return df

@app.post("/unify")
async def unify_files(csv_file: UploadFile = File(...), xlsx_file: UploadFile = File(...)):
    print(f"Receiving files: CSV={csv_file.filename}, Excel={xlsx_file.filename}")
    
    try:
        # Load Mapping Config
        mapping = load_mapping()
        
        # 1. Read files into memory
        csv_content = await csv_file.read()
        xlsx_content = await xlsx_file.read()
        
        # 2. Convert to DataFrames
        try:
            # Read CSV
            df_csv = pd.read_csv(io.BytesIO(csv_content), sep=';', encoding='latin-1', on_bad_lines='skip') # Robust read settings
            if df_csv.shape[1] < 2: # Fallback to comma if semicolon failed
                print("Warning: CSV read with semicolon yielded <2 cols. Retrying with comma.")
                csv_file.file.seek(0)
                df_csv = pd.read_csv(io.BytesIO(csv_content), sep=',')
            
            # Read Excel (Respecting Header Offset)
            # The original requirement stated the header is on row 8 (index 7)
            df_xlsx = pd.read_excel(io.BytesIO(xlsx_content), header=7) 
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Error parsing file: {str(e)}")

        # 3. Identify and Assign Sources
        # ... (skipping unchanged code) ...

        # 4. Transform (Merge)
        
        # Prepare Excel DataFrame
        if 'DATE' in df_xlsx.columns:
            df_xlsx.rename(columns={'DATE': 'DATE_KEY'}, inplace=True)
        elif len(df_xlsx.columns) > 0:
             # Fallback: assumed first column is Date if 'DATE' not found
            df_xlsx.rename(columns={df_xlsx.columns[0]: 'DATE_KEY'}, inplace=True)
            
        df_xlsx['DATE_KEY'] = pd.to_datetime(df_xlsx['DATE_KEY'], dayfirst=True, errors='coerce')
        df_xlsx.dropna(subset=['DATE_KEY'], inplace=True)
        df_xlsx = sanitize_cols(df_xlsx)
        
        # Prepare CSV DataFrame
        if 'Datetime' in df_csv.columns:
            df_csv.rename(columns={'Datetime': 'DATE_KEY'}, inplace=True)
        elif len(df_csv.columns) > 0:
            df_csv.rename(columns={df_csv.columns[0]: 'DATE_KEY'}, inplace=True)
        
        df_csv['DATE_KEY'] = pd.to_datetime(df_csv['DATE_KEY'], dayfirst=False, errors='coerce')
        df_csv.dropna(subset=['DATE_KEY'], inplace=True)
        df_csv = sanitize_cols(df_csv)

        # Merge: Excel (Left) + CSV (Right)
        df_merged = pd.merge(df_xlsx, df_csv, on='DATE_KEY', how='outer', suffixes=('_XLSX', '_CSV'))
        
        # Coalesce Logic
        cols = df_merged.columns
        base_cols = set()
        for c in cols:
            if c.endswith('_CSV'):
                base_cols.add(c[:-4])
            elif c.endswith('_XLSX'):
                base_cols.add(c[:-5])
        
        for base in base_cols:
            col_csv = f"{base}_CSV"
            col_xlsx = f"{base}_XLSX"
            
            if col_csv in df_merged.columns and col_xlsx in df_merged.columns:
                # Combined: CSV priority
                combined = df_merged[col_csv].combine_first(df_merged[col_xlsx])
                # Overwrite Excel column (to keep position left)
                df_merged[col_xlsx] = combined
                df_merged.rename(columns={col_xlsx: base}, inplace=True)
                df_merged.drop(columns=[col_csv], inplace=True)
            elif col_csv in df_merged.columns:
                df_merged.rename(columns={col_csv: base}, inplace=True)
            elif col_xlsx in df_merged.columns:
                df_merged.rename(columns={col_xlsx: base}, inplace=True)

        # Sort
        df_merged.sort_values(by='DATE_KEY', inplace=True)
        
        # Calculate Filename using Excel name + Month/Year of data
        import os
        # User wants to preserve the "Report..." name which likely comes from the Excel template
        base_name = os.path.splitext(xlsx_file.filename)[0]
        
        if not df_merged.empty:
            max_date = df_merged['DATE_KEY'].max()
            suffix = max_date.strftime('%m%Y')
        else:
            from datetime import datetime
            suffix = datetime.now().strftime('%m%Y')
            
        filename = f"{base_name}_{suffix}.xlsx"

        # Export with formatting
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df_merged.to_excel(writer, index=False, sheet_name='Unified_Data', startrow=4)
            
            ws = writer.sheets['Unified_Data']
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            from openpyxl.drawing.image import Image as XLImage
            from datetime import datetime
            import os

            # Styles
            header_fill = PatternFill(start_color="E0F7FA", end_color="E0F7FA", fill_type="solid")
            white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
            ws.sheet_view.showGridLines = False
            thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            bold_font = Font(bold=True)
            title_font = Font(bold=True, size=14, color="006064")

            # Header Background (Full Width for aesthetics, but maybe user wants strict? Let's keep title area wide, but data area strict)
            # Actually, user complained about "format up to EA", so let's stick to max_col check.
            
            # Determine max column index based on Dataframe width
            # Note: startrow=4 means header is at row 5. 
            # df_merged.to_excel writes columns 1 to N.
            max_col_idx = df_merged.shape[1] 

            # Logo & Title Section (Keep simple)
            current_dir = os.getcwd()
            logo_path_jpg = os.path.join(current_dir, "src", "assets", "logo.jpg")
            logo_path_png = os.path.join(current_dir, "src", "assets", "logo.png")
            logo_path = logo_path_jpg if os.path.exists(logo_path_jpg) else (logo_path_png if os.path.exists(logo_path_png) else None)
            
            if logo_path:
                try:
                    img = XLImage(logo_path)
                    img.height = 60
                    img.width = 60
                    ws.add_image(img, 'A1')
                except: pass

            ws['B2'] = "SGA DATA"
            ws['B2'].font = title_font
            ws['B3'] = f"Fecha de emisión: {datetime.now().strftime('%d/%m/%Y')}"
            ws['B3'].font = Font(italic=True, size=10)

            period_str = f"Periodo de Datos: {df_merged['DATE_KEY'].min().strftime('%d/%m/%Y')} - {df_merged['DATE_KEY'].max().strftime('%d/%m/%Y')}" if not df_merged.empty else "Periodo de Datos: N/A"
            ws['B4'] = period_str
            ws['B4'].font = Font(bold=True, size=11, color="006064")

            # Table Styles - STRICT LOOP
            header_row_idx = 5
            
            # 1. Header Row Style
            for col_idx in range(1, max_col_idx + 1):
                cell = ws.cell(row=header_row_idx, column=col_idx)
                cell.fill = header_fill
                cell.font = bold_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = thin_border

            # 2. Data Rows Style
            # ws.max_row includes the data written by to_excel
            for row in range(header_row_idx + 1, ws.max_row + 1):
                for col in range(1, max_col_idx + 1):
                    cell = ws.cell(row=row, column=col)
                    cell.border = thin_border
                    
            # 3. Column Widths
            # Iterate only through used columns
            for i, col in enumerate(ws.columns):
                # Stop if we exceed our data width
                if i >= max_col_idx:
                    break
                    
                max_length = 0
                column_letter = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except: pass
                ws.column_dimensions[column_letter].width = max_length + 2

        output.seek(0)
        
        headers = {
            'Content-Disposition': f'attachment; filename="{filename}"'
        }
        
        return StreamingResponse(output, headers=headers, media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    except Exception as e:
        print(f"Server Error: {e}")
        raise HTTPException(status_code=500, detail=str(e))

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)
