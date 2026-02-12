import pandas as pd
import imaplib
import email
import smtplib
import io
import os
import sys
from datetime import datetime
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.application import MIMEApplication
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# --- CONFIGURACIÓN ---
# GitHub Actions pone las variables en entorno.
# Usamos print() para que salga en los logs de la Action.

IMAP_SERVER = os.getenv('EMAIL_HOST', 'ssl0.ovh.net')
SMTP_HOST = os.getenv('SMTP_HOST', 'ssl0.ovh.net')
SMTP_PORT = int(os.getenv('SMTP_PORT', 587))
EMAIL_USER = os.getenv('EMAIL_USER')
EMAIL_PASS = os.getenv('EMAIL_PASS')

def sanitize_cols(df):
    """Limpia nombres de columnas."""
    df.columns = [str(c).strip().replace(' ', '_').replace('.', '_').upper() for c in df.columns]
    return df

def process_and_reply():
    """Función principal (EJECUCIÓN ÚNICA)."""
    print(f"[{datetime.now()}] Iniciando Job ETL...")

    if not EMAIL_USER or not EMAIL_PASS:
        print("[ERROR] Faltan credenciales EMAIL_USER / EMAIL_PASS. Revisa los Secrets del Repo.")
        sys.exit(1)

    # 1. CONEXIÓN IMAP
    try:
        print(f"[...] Conectando a IMAP {IMAP_SERVER}...")
        mail = imaplib.IMAP4_SSL(IMAP_SERVER)
        mail.login(EMAIL_USER, EMAIL_PASS)
        mail.select('inbox')
        
        # Buscar no leídos
        status, messages = mail.search(None, 'UNSEEN')
        email_ids = messages[0].split()
        
        if not email_ids:
            print("[INFO] No hay correos nuevos. Fin del trabajo.")
            mail.close()
            mail.logout()
            return
            
        print(f"[INFO] {len(email_ids)} correos nuevos encontrados.")
        
    except Exception as e:
        print(f"[ERROR] IMAP: {e}")
        sys.exit(1)

    # 2. CONEXIÓN SMTP (Solo si hay correos)
    try:
        server_smtp = smtplib.SMTP(SMTP_HOST, SMTP_PORT)
        server_smtp.starttls()
        server_smtp.login(EMAIL_USER, EMAIL_PASS)
    except Exception as e:
        print(f"[ERROR] SMTP: {e}")
        try: mail.close(); mail.logout()
        except: pass
        sys.exit(1)

    # 3. PROCESAMIENTO
    processed_count = 0
    for e_id in email_ids:
        try:
            _, msg_data = mail.fetch(e_id, '(RFC822)')
            msg = email.message_from_bytes(msg_data[0][1])
            sender = email.utils.parseaddr(msg['From'])[1]
            subject = msg['Subject']
            
            print(f"  > Procesando email de: {sender} | Asunto: {subject}")
            
            attachments = {}
            for part in msg.walk():
                if part.get_content_maintype() == 'multipart': continue
                if part.get('Content-Disposition') is None: continue
                
                fname = part.get_filename() or ""
                payload = part.get_payload(decode=True)
                
                if fname.lower().endswith('.csv'):
                    attachments['csv'] = payload
                elif fname.lower().endswith('.xlsx'):
                    attachments['xlsx'] = payload
                    attachments['xlsx_name'] = fname

            if 'csv' in attachments and 'xlsx' in attachments:
                # --- ETL START ---
                try:
                    df_csv = pd.read_csv(io.BytesIO(attachments['csv']), sep=';', encoding='latin-1', on_bad_lines='skip')
                    if df_csv.shape[1] < 2: 
                        df_csv = pd.read_csv(io.BytesIO(attachments['csv']), sep=',')
                except: df_csv = pd.DataFrame()

                df_xlsx = pd.read_excel(io.BytesIO(attachments['xlsx']), header=7)

                # Renombrar Keys
                if 'DATE' in df_xlsx.columns: df_xlsx.rename(columns={'DATE': 'DATE_KEY'}, inplace=True)
                elif len(df_xlsx.columns) > 0: df_xlsx.rename(columns={df_xlsx.columns[0]: 'DATE_KEY'}, inplace=True)

                if 'Datetime' in df_csv.columns: df_csv.rename(columns={'Datetime': 'DATE_KEY'}, inplace=True)
                elif len(df_csv.columns) > 0: df_csv.rename(columns={df_csv.columns[0]: 'DATE_KEY'}, inplace=True)

                # Conversión de Fechas
                df_xlsx['DATE_KEY'] = pd.to_datetime(df_xlsx['DATE_KEY'], dayfirst=True, errors='coerce')
                df_xlsx.dropna(subset=['DATE_KEY'], inplace=True)
                df_xlsx = sanitize_cols(df_xlsx)

                df_csv['DATE_KEY'] = pd.to_datetime(df_csv['DATE_KEY'], dayfirst=False, errors='coerce')
                df_csv.dropna(subset=['DATE_KEY'], inplace=True)
                df_csv = sanitize_cols(df_csv)

                # Merge & Coalesce
                df_merged = pd.merge(df_xlsx, df_csv, on='DATE_KEY', how='outer', suffixes=('_XLSX', '_CSV'))
                
                cols = df_merged.columns
                base_cols = set()
                for c in cols:
                    if c.endswith('_CSV'): base_cols.add(c[:-4])
                    if c.endswith('_XLSX'): base_cols.add(c[:-5])

                for base in base_cols:
                    c_csv, c_xlsx = f"{base}_CSV", f"{base}_XLSX"
                    if c_csv in df_merged and c_xlsx in df_merged:
                        df_merged[c_xlsx] = df_merged[c_xlsx].combine_first(df_merged[c_csv])
                        df_merged.rename(columns={c_xlsx: base}, inplace=True)
                        df_merged.drop(columns=[c_csv], inplace=True)
                    elif c_csv in df_merged: df_merged.rename(columns={c_csv: base}, inplace=True)
                    elif c_xlsx in df_merged: df_merged.rename(columns={c_xlsx: base}, inplace=True)

                df_merged.sort_values(by='DATE_KEY', inplace=True)
                
                # Generar Excel
                base_n = os.path.splitext(attachments['xlsx_name'])[0]
                suff = datetime.now().strftime('%m%Y')
                if not df_merged.empty: suff = df_merged['DATE_KEY'].max().strftime('%m%Y')
                out_name = f"{base_n}_{suff}.xlsx"

                output_buffer = io.BytesIO()
                with pd.ExcelWriter(output_buffer, engine='openpyxl') as writer:
                    df_merged.to_excel(writer, index=False, sheet_name='Unified_Data', startrow=4)
                    ws = writer.sheets['Unified_Data']
                    
                    # --- ESTILOS SGA DATA ---
                    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
                    from openpyxl.drawing.image import Image as XLImage

                    # Definiciones
                    header_fill = PatternFill(start_color="E0F7FA", end_color="E0F7FA", fill_type="solid")
                    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                    bold_font = Font(bold=True)
                    title_font = Font(bold=True, size=14, color="006064")
                    ws.sheet_view.showGridLines = False

                    # 1. LOGO
                    # En GitHub Actions, el repo está en la raíz.
                    logo_path = "src/assets/logo.png"
                    if os.path.exists(logo_path):
                        try:
                            img = XLImage(logo_path)
                            img.height = 60
                            img.width = 60
                            ws.add_image(img, 'A1')
                            print("    [STYLE] Logo insertado.")
                        except Exception as e:
                            print(f"    [WARN] No se pudo insertar logo: {e}")

                    # 2. CABECERA
                    ws['B2'] = "SGA DATA"
                    ws['B2'].font = title_font
                    ws['B3'] = f"Fecha de emisión: {datetime.now().strftime('%d/%m/%Y')}"
                    ws['B3'].font = Font(italic=True, size=10)

                    min_date = df_merged['DATE_KEY'].min().strftime('%d/%m/%Y') if not df_merged.empty else "N/A"
                    max_date = df_merged['DATE_KEY'].max().strftime('%d/%m/%Y') if not df_merged.empty else "N/A"
                    ws['B4'] = f"Periodo de Datos: {min_date} - {max_date}"
                    ws['B4'].font = Font(bold=True, size=11, color="006064")

                    # 3. TABLA (Bordes y Cabecera Azul)
                    header_row_idx = 5
                    max_col = df_merged.shape[1]

                    # Estilo Cabecera
                    for col_idx in range(1, max_col + 1):
                        cell = ws.cell(row=header_row_idx, column=col_idx)
                        cell.fill = header_fill
                        cell.font = bold_font
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                        cell.border = thin_border
                    
                    # Estilo Celdas de Datos
                    for row in range(header_row_idx + 1, ws.max_row + 1):
                        for col in range(1, max_col + 1):
                            cell = ws.cell(row=row, column=col)
                            cell.border = thin_border

                    # 4. ANCHO COLUMNAS (Auto-fit simple)
                    for i, col in enumerate(ws.columns):
                        if i >= max_col: break
                        ws.column_dimensions[col[0].column_letter].width = 18

                output_buffer.seek(0)
                final_bytes = output_buffer.getvalue()
                # --- ETL END ---

                # Respuesta SMTP
                msg_out = MIMEMultipart()
                msg_out['From'] = EMAIL_USER
                msg_out['To'] = sender
                msg_out['Subject'] = f"RE: {subject} - Reporte Unificado"
                
                msg_out.attach(MIMEText(f"Procesado correctamente. Archivo: {out_name}", 'plain'))
                part = MIMEApplication(final_bytes, Name=out_name)
                part['Content-Disposition'] = f'attachment; filename="{out_name}"'
                msg_out.attach(part)
                
                server_smtp.send_message(msg_out)
                print(f"    [OK] Respuesta enviada a {sender}")
                processed_count += 1

            else:
                print(f"    [SKIP] Faltan adjuntos CSV/XLSX en el correo.")

        except Exception as e:
             print(f"    [ERROR] Item {e_id}: {e}")

    # Cerrar conexiones
    try:
        mail.close()
        mail.logout()
        server_smtp.quit()
    except: pass
    
    print(f"[FIN] Se procesaron {processed_count} correos.")

if __name__ == "__main__":
    # Sin while True, GitHub Actions manda lanzarlo 1 vez
    process_and_reply()
